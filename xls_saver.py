from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelSaver:
    def __init__(self,file_name):
        self.file_name = file_name
        self.wb = Workbook()
        self.sheet1 = self.wb.active
        self.sheet1.title = 'values_of_consensus'
        self.sheet1.append(['Nodes', 'Percent of Nodes','Percent of Value'])
        self.wb.save(file_name)

    def save_to_file(self,x_values, y_values ):
        wb = load_workbook(self.file_name)
        sheet1_append = wb.active

        self.wb.save(filename=self.file_name)


        for x, y in zip(x_values, y_values):
            sheet1_append.append([x,(x/max(x_values))*100,y])


        wb.save(self.file_name)

class ExcelReader:
    def __init__(self, file_name):
        self.name = file_name

    def read_from_file(self):
        document = load_workbook(self.name, data_only=True)
        sheet = document['values_of_consensus']

        all_values = list()
        for row in sheet.rows:
            row_value = list()
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)

        return all_values.copy()
